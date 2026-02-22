"""CLI interface for Stechuhr."""

import datetime
import platform
import subprocess

import click
from openpyxl import load_workbook

from . import config as cfg_mod
from . import engine, excel

WEEKDAY_NAMES_FULL = ["Montag", "Dienstag", "Mittwoch", "Donnerstag", "Freitag", "Samstag", "Sonntag"]
WEEKDAY_NAMES_SHORT = ["Mo", "Di", "Mi", "Do", "Fr"]


def _fmt_hours(value: float) -> str:
    """Format decimal hours as H:MM. E.g. 2.5 -> '2:30', 0.75 -> '0:45'."""
    sign = "-" if value < 0 else ""
    total_minutes = round(abs(value) * 60)
    h = total_minutes // 60
    m = total_minutes % 60
    return f"{sign}{h}:{m:02d}"


def _color_saldo(value: float, with_sign: bool = True) -> str:
    """Format a saldo value with color: green if positive, red if negative, white if zero."""
    formatted = _fmt_hours(value)
    if with_sign and value >= 0:
        formatted = "+" + formatted
    if value > 0:
        return click.style(formatted, fg="green")
    elif value < 0:
        return click.style(formatted, fg="red")
    return formatted


def _warn_open_stamps(data_dir, config, today):
    """Check if yesterday (or last workday) has an open stamp and warn."""
    check_date = today - datetime.timedelta(days=1)
    # Walk back to find the last workday
    while check_date.weekday() > 4:
        check_date -= datetime.timedelta(days=1)

    wb_path = excel.get_workbook_path(data_dir, check_date.year)
    if not wb_path.exists():
        return

    wb = load_workbook(wb_path)
    sheet_name = excel.MONTH_NAMES[check_date.month - 1]
    if sheet_name not in wb.sheetnames:
        return

    ws = wb[sheet_name]
    if excel.has_open_stamp(ws, check_date):
        click.echo(click.style(
            f"  Achtung: {check_date.strftime('%d.%m.%Y')} ({WEEKDAY_NAMES_FULL[check_date.weekday()]}) "
            f"hat einen offenen Eintrag (Ein ohne Aus)!",
            fg="yellow",
        ))
        click.echo(click.style(
            "  Bitte mit 'stempel nachtrag' korrigieren.",
            fg="yellow",
        ))


@click.group()
def cli():
    """Stechuhr - Zeiterfassung / Time Clock"""
    pass


@cli.command(name="ein")
@click.option("--home", "is_home", is_flag=True, help="Homeoffice (kein Reisezeit-Offset)")
@click.option("--time", "manual_time", type=str, default=None, help="Manuelle Uhrzeit, Format HH:MM")
@click.option("--date", "manual_date", type=str, default=None, help="Manuelles Datum, Format YYYY-MM-DD")
def clock_in(is_home, manual_time, manual_date):
    """Einstempeln (Clock in)"""
    config = cfg_mod.load_config()
    data_dir = cfg_mod.get_data_dir(config)

    now = datetime.datetime.now()
    date = _parse_date(manual_date) if manual_date else now.date()
    stamp_time = _parse_time(manual_time) if manual_time else now.time().replace(second=0, microsecond=0)

    if date.weekday() > 4:
        raise click.ClickException("Wochenende! Keine Zeiterfassung am Samstag/Sonntag.")

    wb = excel.ensure_workbook(data_dir, date.year, config)
    sheet_name = excel.MONTH_NAMES[date.month - 1]
    ws = wb[sheet_name]

    excel.write_stamp(ws, date, stamp_time, "ein", is_home, config)
    excel.save_workbook(wb, data_dir, date.year)

    weekday = WEEKDAY_NAMES_FULL[date.weekday()]
    modus_text = " (Homeoffice)" if is_home else ""

    click.echo(f"Eingestempelt um {stamp_time.strftime('%H:%M')}{modus_text}")
    click.echo(f"Datum: {date.strftime('%d.%m.%Y')} ({weekday})")

    # Warn if last workday has an open stamp
    _warn_open_stamps(data_dir, config, date)


@cli.command(name="aus")
@click.option("--krank", "is_sick", is_flag=True, help="Krank: Soll-Stunden werden als gearbeitet gezaehlt")
@click.option("--time", "manual_time", type=str, default=None, help="Manuelle Uhrzeit, Format HH:MM")
@click.option("--date", "manual_date", type=str, default=None, help="Manuelles Datum, Format YYYY-MM-DD")
def clock_out(is_sick, manual_time, manual_date):
    """Ausstempeln (Clock out)"""
    config = cfg_mod.load_config()
    data_dir = cfg_mod.get_data_dir(config)

    now = datetime.datetime.now()
    date = _parse_date(manual_date) if manual_date else now.date()
    stamp_time = _parse_time(manual_time) if manual_time else now.time().replace(second=0, microsecond=0)

    if date.weekday() > 4:
        raise click.ClickException("Wochenende! Keine Zeiterfassung am Samstag/Sonntag.")

    wb = excel.ensure_workbook(data_dir, date.year, config)
    sheet_name = excel.MONTH_NAMES[date.month - 1]
    ws = wb[sheet_name]

    try:
        excel.write_stamp(ws, date, stamp_time, "aus", False, config)
    except ValueError as e:
        raise click.ClickException(str(e))

    # If sick, override Gesamt to Soll so no hours are lost
    if is_sick:
        excel.set_day_sick(ws, date, config)

    excel.save_workbook(wb, data_dir, date.year)

    weekday = WEEKDAY_NAMES_FULL[date.weekday()]
    day_row = excel.read_day_row(ws, date, config)

    modus_text = ""
    if is_sick:
        modus_text = " (krank)"

    click.echo(f"Ausgestempelt um {stamp_time.strftime('%H:%M')}{modus_text}")
    click.echo(f"Datum: {date.strftime('%d.%m.%Y')} ({weekday})")

    if is_sick:
        click.echo(click.style("Gute Besserung! Soll-Stunden werden als gearbeitet gezaehlt.", fg="yellow"))
    elif day_row and day_row.total is not None:
        soll = day_row.expected
        saldo = day_row.total - soll
        click.echo(f"Gearbeitet heute: {_fmt_hours(day_row.total)} | Soll: {_fmt_hours(soll)} | Saldo: {_color_saldo(saldo)}")


@cli.command(name="saldo")
def overtime():
    """Saldo anzeigen (kumulativ bis gestern)"""
    config = cfg_mod.load_config()
    data_dir = cfg_mod.get_data_dir(config)

    today = datetime.date.today()
    balance = engine.calculate_overtime_balance(data_dir, config, today)

    yesterday = today - datetime.timedelta(days=1)
    click.echo(f"Saldo bis {yesterday.strftime('%d.%m.%Y')}: {_color_saldo(balance)}")


@cli.command(name="status")
def status():
    """Heutigen Status anzeigen"""
    config = cfg_mod.load_config()
    data_dir = cfg_mod.get_data_dir(config)

    today = datetime.date.today()
    weekday = WEEKDAY_NAMES_FULL[today.weekday()]

    day_row = engine.get_today_status(data_dir, config, today)

    click.echo(f"Status fuer {today.strftime('%d.%m.%Y')} ({weekday}):")

    if day_row is None or not day_row.blocks:
        click.echo("  Keine Eintraege fuer heute.")
        return

    for i, block in enumerate(day_row.blocks, start=1):
        ein_str = block.ein.strftime("%H:%M") if block.ein else "--:--"
        aus_str = block.aus.strftime("%H:%M") if block.aus else "--:--"
        hours_str = _fmt_hours(block.hours) if block.hours is not None else "--"
        click.echo(f"  Ein {i}: {ein_str}   Aus {i}: {aus_str}   Stunden: {hours_str}")

    if day_row.status:
        click.echo(f"  Status: {day_row.status}")

    # Check if currently clocked in (open stamp)
    has_open = any(b.ein is not None and b.aus is None for b in day_row.blocks)

    if has_open:
        # Calculate live hours including current time
        current_hours = engine.get_current_hours(data_dir, config, today)
        if current_hours is not None:
            soll = day_row.expected
            saldo = current_hours - soll
            click.echo(f"  Gesamt: ~{_fmt_hours(current_hours)} (laufend) | Soll: {_fmt_hours(soll)} | Saldo: ~{_color_saldo(saldo)}")
        else:
            click.echo("  Gesamt: -- (offener Eintrag)")
    elif day_row.total is not None:
        soll = day_row.expected
        saldo = day_row.total - soll
        click.echo(f"  Gesamt: {_fmt_hours(day_row.total)} | Soll: {_fmt_hours(soll)} | Saldo: {_color_saldo(saldo)}")
    else:
        click.echo("  Gesamt: -- (offener Eintrag)")


@cli.command(name="woche")
def week_summary():
    """Wochenuebersicht anzeigen"""
    config = cfg_mod.load_config()
    data_dir = cfg_mod.get_data_dir(config)

    today = datetime.date.today()
    # Find Monday of this week
    monday = today - datetime.timedelta(days=today.weekday())

    click.echo(f"Woche {monday.strftime('%d.%m.')} - {(monday + datetime.timedelta(days=4)).strftime('%d.%m.%Y')}:")
    click.echo()

    week_total = 0.0
    week_soll = 0.0

    for i in range(5):  # Mon-Fri
        day = monday + datetime.timedelta(days=i)
        day_row = engine.get_today_status(data_dir, config, day)

        tag = WEEKDAY_NAMES_SHORT[i]
        date_str = day.strftime("%d.%m.")
        marker = " <--" if day == today else ""

        if day_row and day_row.total is not None:
            # Check if this day has an open stamp (today, currently clocked in)
            has_open = any(b.ein is not None and b.aus is None for b in day_row.blocks)
            if has_open and day == today:
                current_hours = engine.get_current_hours(data_dir, config, today)
                if current_hours is not None:
                    soll = day_row.expected
                    saldo = current_hours - soll
                    week_total += current_hours
                    week_soll += soll
                    click.echo(
                        f"  {tag} {date_str}  "
                        f"Gesamt: ~{_fmt_hours(current_hours):>4}  "
                        f"Soll: {_fmt_hours(soll):>5}  "
                        f"Saldo: ~{_color_saldo(saldo)}"
                        f"{click.style(marker, fg='cyan')}"
                    )
                    continue

            soll = day_row.expected
            saldo = day_row.total - soll
            week_total += day_row.total
            week_soll += soll
            click.echo(
                f"  {tag} {date_str}  "
                f"Gesamt: {_fmt_hours(day_row.total):>5}  "
                f"Soll: {_fmt_hours(soll):>5}  "
                f"Saldo: {_color_saldo(saldo)}"
                f"{click.style(marker, fg='cyan')}"
            )
        elif day_row and day_row.blocks:
            # Has stamps but no total yet (open entry)
            has_open = any(b.ein is not None and b.aus is None for b in day_row.blocks)
            soll = day_row.expected
            week_soll += soll

            if has_open and day == today:
                current_hours = engine.get_current_hours(data_dir, config, today)
                if current_hours is not None:
                    saldo = current_hours - soll
                    week_total += current_hours
                    click.echo(
                        f"  {tag} {date_str}  "
                        f"Gesamt: ~{_fmt_hours(current_hours):>4}  "
                        f"Soll: {_fmt_hours(soll):>5}  "
                        f"Saldo: ~{_color_saldo(saldo)}"
                        f"{click.style(marker, fg='cyan')}"
                    )
                    continue

            click.echo(
                f"  {tag} {date_str}  "
                f"Gesamt:    --  "
                f"Soll: {_fmt_hours(soll):>5}  "
                f"Saldo:   --"
                f"{click.style(marker, fg='cyan')}"
            )
        else:
            soll = cfg_mod.get_expected_hours(config, day.weekday())
            week_soll += soll
            if day <= today:
                click.echo(
                    f"  {tag} {date_str}  "
                    f"Gesamt:    --  "
                    f"Soll: {_fmt_hours(soll):>5}  "
                    f"Saldo:   --"
                    f"{click.style(marker, fg='cyan')}"
                )
            else:
                click.echo(
                    f"  {tag} {date_str}  "
                    f"{click.style('(noch offen)', dim=True)}"
                )

    click.echo()
    week_saldo = week_total - week_soll
    click.echo(
        f"  Woche:     "
        f"Gesamt: {_fmt_hours(week_total):>5}  "
        f"Soll: {_fmt_hours(week_soll):>5}  "
        f"Saldo: {_color_saldo(week_saldo)}"
    )


@cli.command(name="excel")
def open_file():
    """Excel-Datei im Standardprogramm oeffnen"""
    config = cfg_mod.load_config()
    data_dir = cfg_mod.get_data_dir(config)

    year = datetime.date.today().year
    path = excel.get_workbook_path(data_dir, year)

    if not path.exists():
        raise click.ClickException(f"Datei nicht gefunden: {path}")

    system = platform.system()
    if system == "Darwin":
        subprocess.Popen(["open", str(path)])
    elif system == "Windows":
        subprocess.Popen(["start", "", str(path)], shell=True)
    else:
        subprocess.Popen(["xdg-open", str(path)])

    click.echo(f"Oeffne {path.name}...")


@cli.command(name="nachtrag")
@click.option("--date", "target_date", required=True, type=str, help="Datum, Format YYYY-MM-DD")
@click.option("--ein", "ein_time", type=str, default=None, help="Einstempelzeit HH:MM")
@click.option("--aus", "aus_time", type=str, default=None, help="Ausstempelzeit HH:MM")
@click.option("--home", "is_home", is_flag=True, help="Homeoffice")
def manual_entry(target_date, ein_time, aus_time, is_home):
    """Nachtrag: Manuellen Zeiteintrag hinzufuegen"""
    if not ein_time and not aus_time:
        raise click.ClickException("Mindestens --ein oder --aus angeben.")

    if ein_time and aus_time:
        ein_t = _parse_time(ein_time)
        aus_t = _parse_time(aus_time)
        if aus_t <= ein_t:
            raise click.ClickException(f"Aus-Zeit ({aus_time}) muss nach Ein-Zeit ({ein_time}) liegen.")

    config = cfg_mod.load_config()
    data_dir = cfg_mod.get_data_dir(config)

    date = _parse_date(target_date)

    if date.weekday() > 4:
        raise click.ClickException("Wochenende! Keine Zeiterfassung am Samstag/Sonntag.")

    wb = excel.ensure_workbook(data_dir, date.year, config)
    sheet_name = excel.MONTH_NAMES[date.month - 1]
    ws = wb[sheet_name]

    if ein_time:
        t = _parse_time(ein_time)
        excel.write_stamp(ws, date, t, "ein", is_home, config)

    if aus_time:
        t = _parse_time(aus_time)
        try:
            excel.write_stamp(ws, date, t, "aus", False, config)
        except ValueError as e:
            raise click.ClickException(str(e))

    excel.save_workbook(wb, data_dir, date.year)

    click.echo(f"Nachtrag fuer {date.strftime('%d.%m.%Y')} gespeichert.")
    if ein_time:
        click.echo(f"  Ein: {ein_time}")
    if aus_time:
        click.echo(f"  Aus: {aus_time}")
    if is_home:
        click.echo(f"  Status: Home")


@cli.command(name="update")
@click.option("--date", "target_date", type=str, default=None,
              help="Einzelnes Datum neu berechnen, Format YYYY-MM-DD")
@click.option("--month", "target_month", type=str, default=None,
              help="Ganzen Monat neu berechnen, Format YYYY-MM")
def update(target_date, target_month):
    """Neuberechnung nach manueller Excel-Bearbeitung"""
    config = cfg_mod.load_config()
    data_dir = cfg_mod.get_data_dir(config)

    if target_date:
        date = _parse_date(target_date)
        wb = excel.ensure_workbook(data_dir, date.year, config)
        ws = wb[excel.MONTH_NAMES[date.month - 1]]
        excel.recalculate_day(ws, date, config)
        excel.save_workbook(wb, data_dir, date.year)
        click.echo(f"Neuberechnung fuer {date.strftime('%d.%m.%Y')} abgeschlossen.")

    elif target_month:
        try:
            parts = target_month.strip().split("-")
            year, month = int(parts[0]), int(parts[1])
        except (ValueError, IndexError):
            raise click.ClickException(f"Ungueltiger Monat: '{target_month}'. Format: YYYY-MM")

        wb = excel.ensure_workbook(data_dir, year, config)
        sheet_name = excel.MONTH_NAMES[month - 1]
        ws = wb[sheet_name]

        count = 0
        for row in range(excel.HEADER_ROW + 1, ws.max_row + 1):
            date_val = ws.cell(row, 1).value
            if date_val is None:
                continue
            if isinstance(date_val, str) and date_val in ("Summe", "Uebertrag", "Kumuliert"):
                break
            if isinstance(date_val, str):
                try:
                    dt = datetime.datetime.strptime(date_val, "%d.%m.%Y").date()
                except ValueError:
                    continue
            else:
                continue

            num_blocks = excel._count_blocks(ws)
            has_stamps = any(
                ws.cell(row, excel._ein_col(b)).value is not None
                for b in range(1, num_blocks + 1)
            )
            if has_stamps:
                excel.recalculate_day(ws, dt, config)
                count += 1

        excel.save_workbook(wb, data_dir, year)
        click.echo(f"Neuberechnung fuer {sheet_name} {year}: {count} Tage aktualisiert.")

    else:
        year = datetime.date.today().year
        wb = excel.ensure_workbook(data_dir, year, config)
        total_count = 0

        for month in range(1, 13):
            sheet_name = excel.MONTH_NAMES[month - 1]
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]

            for row in range(excel.HEADER_ROW + 1, ws.max_row + 1):
                date_val = ws.cell(row, 1).value
                if date_val is None:
                    continue
                if isinstance(date_val, str) and date_val in ("Summe", "Uebertrag", "Kumuliert"):
                    break
                if isinstance(date_val, str):
                    try:
                        dt = datetime.datetime.strptime(date_val, "%d.%m.%Y").date()
                    except ValueError:
                        continue
                else:
                    continue

                num_blocks = excel._count_blocks(ws)
                has_stamps = any(
                    ws.cell(row, excel._ein_col(b)).value is not None
                    for b in range(1, num_blocks + 1)
                )
                if has_stamps:
                    excel.recalculate_day(ws, dt, config)
                    total_count += 1

        excel.save_workbook(wb, data_dir, year)
        click.echo(f"Neuberechnung fuer {year}: {total_count} Tage aktualisiert.")


@cli.command(name="config")
def show_config():
    """Aktuelle Konfiguration anzeigen"""
    config = cfg_mod.load_config()
    click.echo("Stechuhr Konfiguration:")
    click.echo(f"  Konfig-Datei: {cfg_mod.get_config_path()}")
    click.echo(f"  Daten-Verzeichnis: {config['data_dir']}")
    click.echo(f"  Reisezeit-Offset: {config['travel_offset_minutes']} Minuten")
    click.echo("  Soll-Stunden:")
    for day, hours in config["expected_hours"].items():
        click.echo(f"    {day}: {_fmt_hours(hours)}")
    click.echo(f"  Pausen-Schwelle: {_fmt_hours(config['auto_break_threshold_hours'])}")
    click.echo(f"  Pausen-Abzug: {config['auto_break_deduction_minutes']} Minuten")
    if config.get("carry_over_balance"):
        click.echo("  Uebertrag:")
        for year, bal in config["carry_over_balance"].items():
            click.echo(f"    {year}: {_color_saldo(bal)}")


def _parse_time(time_str: str) -> datetime.time:
    """Parse HH:MM string to time."""
    try:
        parts = time_str.strip().split(":")
        return datetime.time(int(parts[0]), int(parts[1]))
    except (ValueError, IndexError):
        raise click.ClickException(f"Ungueltige Uhrzeit: '{time_str}'. Format: HH:MM")


def _parse_date(date_str: str) -> datetime.date:
    """Parse YYYY-MM-DD string to date."""
    try:
        return datetime.datetime.strptime(date_str.strip(), "%Y-%m-%d").date()
    except ValueError:
        raise click.ClickException(f"Ungueltiges Datum: '{date_str}'. Format: YYYY-MM-DD")


if __name__ == "__main__":
    cli()
