"""Core business logic for Stechuhr: overtime calculation, missing day fill."""

import datetime
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from . import config as cfg_mod
from . import excel


def calculate_overtime_balance(data_dir: Path, config: dict,
                               up_to_date: datetime.date) -> float:
    """Calculate cumulative overtime from start of tracking to up_to_date (exclusive).

    Walks all years/months, fills missing days, sums saldo per day.
    """
    total_balance = 0.0

    # Find all year files
    year_files = sorted(data_dir.glob("*.xlsx"))
    if not year_files:
        return 0.0

    earliest_year = None
    for yf in year_files:
        try:
            y = int(yf.stem)
            if earliest_year is None or y < earliest_year:
                earliest_year = y
        except ValueError:
            continue

    if earliest_year is None:
        return 0.0

    # Add carry-over from before the earliest year
    for year_str, balance in config.get("carry_over_balance", {}).items():
        try:
            y = int(year_str)
            if y < earliest_year:
                total_balance += balance
        except ValueError:
            continue

    # Process each year
    for year in range(earliest_year, up_to_date.year + 1):
        wb_path = excel.get_workbook_path(data_dir, year)
        if not wb_path.exists():
            continue

        wb = load_workbook(wb_path)

        # Determine carry-over for this year
        year_carry = cfg_mod.get_carry_over(config, year)
        month_carry = year_carry

        for month in range(1, 13):
            if year == up_to_date.year and month > up_to_date.month:
                break

            sheet_name = excel.MONTH_NAMES[month - 1]
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]

            # Fill missing days up to the cutoff
            excel.fill_missing_days(ws, config, up_to_date)

            # Update sheet summary and get cumulative balance
            month_carry = excel.recalculate_sheet_summary(
                ws, year, month, month_carry, config
            )

            # Read all day rows for total_balance calculation
            day_rows = excel.iter_day_rows_with_data(ws)
            for dt, total, expected, saldo in day_rows:
                if dt >= up_to_date:
                    continue

                if saldo is not None:
                    total_balance += saldo
                elif total is not None:
                    total_balance += (total - expected)

        # Save the workbook with filled missing days and updated summaries
        excel.save_workbook(wb, data_dir, year)

    return round(total_balance, 2)


def get_today_status(data_dir: Path, config: dict,
                     date: datetime.date) -> Optional[excel.DayRow]:
    """Read today's row from the workbook."""
    wb_path = excel.get_workbook_path(data_dir, date.year)
    if not wb_path.exists():
        return None

    wb = load_workbook(wb_path)
    sheet_name = excel.MONTH_NAMES[date.month - 1]
    if sheet_name not in wb.sheetnames:
        return None

    ws = wb[sheet_name]
    return excel.read_day_row(ws, date, config)


def get_current_hours(data_dir: Path, config: dict,
                      date: datetime.date) -> Optional[float]:
    """Calculate hours worked so far today, treating 'now' as a virtual clock-out."""
    wb_path = excel.get_workbook_path(data_dir, date.year)
    if not wb_path.exists():
        return None

    wb = load_workbook(wb_path)
    sheet_name = excel.MONTH_NAMES[date.month - 1]
    if sheet_name not in wb.sheetnames:
        return None

    ws = wb[sheet_name]
    return excel.calculate_current_hours(ws, date, config)
