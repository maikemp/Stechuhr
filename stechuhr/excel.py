"""Excel workbook management for Stechuhr."""

import calendar
import datetime
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from . import config as cfg_mod

MONTH_NAMES = [
    "Jan", "Feb", "Mar", "Apr", "Mai", "Jun",
    "Jul", "Aug", "Sep", "Okt", "Nov", "Dez",
]

WEEKDAY_NAMES = ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"]

# Excel number formats for H:MM display
HOURS_FMT = "[h]:mm"
SALDO_FMT = '[h]:mm;\\-[h]:mm;"0:00"'

# Column layout constants
# A=Datum, B=Tag, then blocks of 3 (Ein, Aus, Stunden), then Status, Gesamt, Soll, Saldo
DATE_COLS = 2  # Datum + Tag
BLOCK_WIDTH = 3  # Ein, Aus, Stunden (no more Modus per block)
INITIAL_BLOCKS = 3
HEADER_ROW = 1


@dataclass
class StampBlock:
    ein: Optional[datetime.time] = None
    aus: Optional[datetime.time] = None
    hours: Optional[float] = None


@dataclass
class DayRow:
    date: datetime.date = None
    weekday_name: str = ""
    blocks: List[StampBlock] = field(default_factory=list)
    status: str = ""  # "Home" or "Office" or ""
    total: Optional[float] = None
    expected: float = 0.0
    balance: Optional[float] = None
    row_num: int = 0


def _hours_to_fraction(hours: float) -> float:
    """Convert decimal hours to Excel time fraction (hours/24)."""
    return hours / 24.0


def _fraction_to_hours(fraction: float) -> float:
    """Convert Excel time fraction back to decimal hours."""
    return fraction * 24.0


def _read_hours_value(val) -> float:
    """Read an hours value from a cell, handling timedelta, time-fraction and legacy decimal formats.

    openpyxl reads cells with [h]:mm format as datetime.timedelta objects.
    Time fractions: 8h = 8/24 = 0.333. Legacy decimals: 6.0, 8.0, etc.
    Heuristic for floats: if abs(val) < 1.0, treat as time fraction; otherwise legacy decimal.
    After running 'stempel update', all values will be time fractions (read back as timedelta).
    """
    if val is None:
        return 0.0
    if isinstance(val, datetime.timedelta):
        return round(val.total_seconds() / 3600.0, 2)
    if isinstance(val, datetime.time):
        # Should not normally happen, but guard against it
        return round(val.hour + val.minute / 60.0, 2)
    v = float(val)
    if abs(v) < 1.0:
        return round(_fraction_to_hours(v), 2)
    return v


def _ein_col(block_num: int) -> int:
    """1-based column for Ein of block N (block_num starts at 1)."""
    return DATE_COLS + (block_num - 1) * BLOCK_WIDTH + 1


def _aus_col(block_num: int) -> int:
    return _ein_col(block_num) + 1


def _stunden_col(block_num: int) -> int:
    return _ein_col(block_num) + 2


def _find_summary_cols(ws: Worksheet) -> Tuple[int, int, int, int]:
    """Find Status, Gesamt, Soll, Saldo column indices by scanning header row.
    Returns (status_col, gesamt_col, soll_col, saldo_col).
    """
    for col in range(1, ws.max_column + 1):
        val = ws.cell(HEADER_ROW, col).value
        if val == "Status":
            return col, col + 1, col + 2, col + 3
    raise ValueError("Summary columns (Status/Gesamt/Soll/Saldo) not found in header row")


def _count_blocks(ws: Worksheet) -> int:
    """Count how many stamp blocks exist by scanning header row."""
    count = 0
    col = DATE_COLS + 1
    while col <= ws.max_column:
        val = ws.cell(HEADER_ROW, col).value
        if val and str(val).startswith("Ein"):
            count += 1
            col += BLOCK_WIDTH
        else:
            break
    return max(count, INITIAL_BLOCKS)


def get_workbook_path(data_dir: Path, year: int) -> Path:
    return data_dir / f"{year}.xlsx"


def ensure_workbook(data_dir: Path, year: int, config: dict) -> Workbook:
    """Load existing or create new workbook for the given year."""
    path = get_workbook_path(data_dir, year)
    if path.exists():
        return load_workbook(path)

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for month in range(1, 13):
        _create_month_sheet(wb, year, month, config)

    save_workbook(wb, data_dir, year)
    return wb


def _create_month_sheet(wb: Workbook, year: int, month: int, config: dict) -> Worksheet:
    """Create a sheet for a given month with headers and workday rows."""
    ws = wb.create_sheet(title=MONTH_NAMES[month - 1])

    # Styling
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write headers
    headers = ["Datum", "Tag"]
    for b in range(1, INITIAL_BLOCKS + 1):
        headers.extend([f"Ein {b}", f"Aus {b}", f"Stunden {b}"])
    headers.extend(["Status", "Gesamt", "Soll", "Saldo"])

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(HEADER_ROW, col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # Set column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 5

    # Write workday rows
    row = HEADER_ROW + 1
    cal = calendar.Calendar()
    for day in cal.itermonthdays2(year, month):
        day_num, weekday = day
        if day_num == 0:
            continue  # day belongs to another month
        # Only create rows for days with expected hours > 0
        expected = cfg_mod.get_expected_hours(config, weekday)
        if expected <= 0:
            continue

        dt = datetime.date(year, month, day_num)
        ws.cell(row, 1, value=dt.strftime("%d.%m.%Y"))
        ws.cell(row, 1).alignment = Alignment(horizontal="center")
        ws.cell(row, 1).border = thin_border

        ws.cell(row, 2, value=WEEKDAY_NAMES[weekday])
        ws.cell(row, 2).alignment = Alignment(horizontal="center")
        ws.cell(row, 2).border = thin_border

        # Set expected hours (Soll) - stored as time fraction for [h]:mm display
        status_col, gesamt_col, soll_col, saldo_col = _get_summary_col_indices(INITIAL_BLOCKS)
        expected = cfg_mod.get_expected_hours(config, weekday)
        ws.cell(row, soll_col, value=_hours_to_fraction(expected))
        ws.cell(row, soll_col).number_format = HOURS_FMT
        ws.cell(row, soll_col).alignment = Alignment(horizontal="center")
        ws.cell(row, soll_col).border = thin_border

        # Apply borders to all stamp cells
        for col_idx in range(3, saldo_col + 1):
            ws.cell(row, col_idx).border = thin_border
            ws.cell(row, col_idx).alignment = Alignment(horizontal="center")

        row += 1

    # Summary rows
    last_data_row = row - 1
    summary_start = row + 1  # leave a blank row

    status_col, gesamt_col, soll_col, saldo_col = _get_summary_col_indices(INITIAL_BLOCKS)

    # Summe row
    sum_row = summary_start
    ws.cell(sum_row, 1, value="Summe")
    ws.cell(sum_row, 1).font = Font(bold=True)
    ws.cell(sum_row, 1).border = thin_border

    for c in [gesamt_col, soll_col]:
        cell = ws.cell(sum_row, c)
        cell.border = thin_border
        cell.font = Font(bold=True)
        cell.number_format = HOURS_FMT
        cell.alignment = Alignment(horizontal="center")
    saldo_cell = ws.cell(sum_row, saldo_col)
    saldo_cell.border = thin_border
    saldo_cell.font = Font(bold=True)
    saldo_cell.number_format = SALDO_FMT
    saldo_cell.alignment = Alignment(horizontal="center")

    # Uebertrag row
    carry_row = summary_start + 1
    ws.cell(carry_row, 1, value="Uebertrag")
    ws.cell(carry_row, 1).font = Font(bold=True)
    ws.cell(carry_row, 1).border = thin_border
    ws.cell(carry_row, saldo_col).border = thin_border
    ws.cell(carry_row, saldo_col).number_format = SALDO_FMT
    ws.cell(carry_row, saldo_col).alignment = Alignment(horizontal="center")

    # Kumuliert row
    cum_row = summary_start + 2
    ws.cell(cum_row, 1, value="Kumuliert")
    ws.cell(cum_row, 1).font = Font(bold=True, color="FF0000")
    ws.cell(cum_row, 1).border = thin_border
    ws.cell(cum_row, saldo_col).border = thin_border
    ws.cell(cum_row, saldo_col).font = Font(bold=True)
    ws.cell(cum_row, saldo_col).number_format = SALDO_FMT
    ws.cell(cum_row, saldo_col).alignment = Alignment(horizontal="center")

    return ws


def _get_summary_col_indices(num_blocks: int) -> Tuple[int, int, int, int]:
    """Return (status_col, gesamt_col, soll_col, saldo_col) given the number of blocks."""
    status = DATE_COLS + num_blocks * BLOCK_WIDTH + 1
    return status, status + 1, status + 2, status + 3


def find_day_row(ws: Worksheet, date: datetime.date) -> Optional[int]:
    """Find the row number for a given date. Returns None if not found."""
    date_str = date.strftime("%d.%m.%Y")
    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        val = ws.cell(row, 1).value
        if val == date_str:
            return row
        # Also handle if stored as datetime
        if isinstance(val, datetime.datetime) and val.date() == date:
            return row
    return None


def read_day_row(ws: Worksheet, date: datetime.date, config: dict) -> Optional[DayRow]:
    """Read all stamp data for a given date."""
    row_num = find_day_row(ws, date)
    if row_num is None:
        return None

    num_blocks = _count_blocks(ws)
    blocks = []
    for b in range(1, num_blocks + 1):
        ein_val = ws.cell(row_num, _ein_col(b)).value
        aus_val = ws.cell(row_num, _aus_col(b)).value
        stunden_val = ws.cell(row_num, _stunden_col(b)).value

        # Convert datetime to time if needed
        ein_time = _to_time(ein_val)
        aus_time = _to_time(aus_val)

        if ein_time is not None or aus_time is not None:
            hours_val = None
            if stunden_val is not None:
                hours_val = _read_hours_value(stunden_val)
            blocks.append(StampBlock(
                ein=ein_time,
                aus=aus_time,
                hours=hours_val,
            ))

    status_col, gesamt_col, soll_col, saldo_col = _find_summary_cols(ws)
    status_val = ws.cell(row_num, status_col).value or ""
    total_val = ws.cell(row_num, gesamt_col).value
    soll_val = ws.cell(row_num, soll_col).value
    saldo_val = ws.cell(row_num, saldo_col).value

    weekday_val = ws.cell(row_num, 2).value or ""

    return DayRow(
        date=date,
        weekday_name=str(weekday_val),
        blocks=blocks,
        status=str(status_val),
        total=_read_hours_value(total_val) if total_val is not None else None,
        expected=_read_hours_value(soll_val) if soll_val is not None else 0.0,
        balance=_read_hours_value(saldo_val) if saldo_val is not None else None,
        row_num=row_num,
    )


def _to_time(val) -> Optional[datetime.time]:
    """Convert various cell value types to datetime.time."""
    if val is None:
        return None
    if isinstance(val, datetime.time):
        return val
    if isinstance(val, datetime.datetime):
        return val.time()
    if isinstance(val, str):
        val = val.strip()
        if not val:
            return None
        try:
            parts = val.split(":")
            return datetime.time(int(parts[0]), int(parts[1]))
        except (ValueError, IndexError):
            return None
    return None


def _get_status_col(ws: Worksheet) -> int:
    """Find the Status column index."""
    status_col, _, _, _ = _find_summary_cols(ws)
    return status_col


def _read_day_status(ws: Worksheet, row_num: int) -> str:
    """Read the Status value for a day row."""
    status_col = _get_status_col(ws)
    if status_col is None:
        return ""
    val = ws.cell(row_num, status_col).value
    return str(val) if val else ""


def _insert_day_row(ws: Worksheet, date: datetime.date, config: dict) -> int:
    """Insert a new row for a date that doesn't exist yet (e.g. weekends).

    Returns the row number of the newly inserted row.
    """
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Find the correct position to insert (maintain date order)
    insert_row = None
    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        date_val = ws.cell(row, 1).value
        if date_val is None:
            continue
        # Stop at summary rows
        if isinstance(date_val, str) and date_val in ("Summe", "Uebertrag", "Kumuliert"):
            insert_row = row
            break
        # Parse existing date
        if isinstance(date_val, str):
            try:
                dt = datetime.datetime.strptime(date_val, "%d.%m.%Y").date()
                if dt > date:
                    insert_row = row
                    break
            except ValueError:
                continue

    if insert_row is None:
        # Shouldn't happen, but fallback to after last data row
        insert_row = ws.max_row + 1

    # Insert new row
    ws.insert_rows(insert_row)

    # Set up the row
    num_blocks = _count_blocks(ws)
    status_col, gesamt_col, soll_col, saldo_col = _find_summary_cols(ws)

    # Date column
    ws.cell(insert_row, 1, value=date.strftime("%d.%m.%Y"))
    ws.cell(insert_row, 1).alignment = Alignment(horizontal="center")
    ws.cell(insert_row, 1).border = thin_border

    # Weekday column
    ws.cell(insert_row, 2, value=WEEKDAY_NAMES[date.weekday()])
    ws.cell(insert_row, 2).alignment = Alignment(horizontal="center")
    ws.cell(insert_row, 2).border = thin_border

    # Expected hours (Soll) - 0 for days without expected hours
    expected = cfg_mod.get_expected_hours(config, date.weekday())
    ws.cell(insert_row, soll_col, value=_hours_to_fraction(expected))
    ws.cell(insert_row, soll_col).number_format = HOURS_FMT
    ws.cell(insert_row, soll_col).alignment = Alignment(horizontal="center")
    ws.cell(insert_row, soll_col).border = thin_border

    # Apply borders to all cells
    for col_idx in range(3, saldo_col + 1):
        ws.cell(insert_row, col_idx).border = thin_border
        ws.cell(insert_row, col_idx).alignment = Alignment(horizontal="center")

    return insert_row


def write_stamp(ws: Worksheet, date: datetime.date, stamp_time: datetime.time,
                stamp_type: str, is_home: bool, config: dict) -> None:
    """Write a clock-in or clock-out stamp to the correct cell.

    stamp_type: "ein" or "aus"
    is_home: True if home office (no travel offset for the day)
    """
    row_num = find_day_row(ws, date)
    if row_num is None:
        # Day doesn't have a row yet (e.g. weekend) - create it
        row_num = _insert_day_row(ws, date, config)

    num_blocks = _count_blocks(ws)

    if stamp_type == "ein":
        # Find next empty Ein column
        target_block = None
        for b in range(1, num_blocks + 1):
            if ws.cell(row_num, _ein_col(b)).value is None:
                target_block = b
                break

        if target_block is None:
            # Need to expand: add a new block
            target_block = num_blocks + 1
            _expand_blocks(ws, target_block)

        ws.cell(row_num, _ein_col(target_block), value=stamp_time)
        ws.cell(row_num, _ein_col(target_block)).number_format = "HH:MM"
        ws.cell(row_num, _ein_col(target_block)).alignment = Alignment(horizontal="center")

        # On first clock-in, set the day's status
        if target_block == 1:
            status_col = _get_status_col(ws)
            status_text = "Home" if is_home else "Office"
            ws.cell(row_num, status_col, value=status_text)
            ws.cell(row_num, status_col).alignment = Alignment(horizontal="center")

    elif stamp_type == "aus":
        # Find the last block that has an Ein but no Aus
        target_block = None
        for b in range(num_blocks, 0, -1):
            ein_val = ws.cell(row_num, _ein_col(b)).value
            aus_val = ws.cell(row_num, _aus_col(b)).value
            if ein_val is not None and aus_val is None:
                target_block = b
                break

        if target_block is None:
            raise ValueError(
                "Kein offener Eintrag gefunden. Bitte zuerst einstempeln."
            )

        ws.cell(row_num, _aus_col(target_block), value=stamp_time)
        ws.cell(row_num, _aus_col(target_block)).number_format = "HH:MM"
        ws.cell(row_num, _aus_col(target_block)).alignment = Alignment(horizontal="center")

    # Recalculate after writing
    recalculate_day(ws, date, config)


def _expand_blocks(ws: Worksheet, new_block_num: int) -> None:
    """Add columns for a new block before the summary columns."""
    insert_col = _get_status_col(ws)

    # Insert 3 columns at the insert position (before Status/Gesamt)
    ws.insert_cols(insert_col, BLOCK_WIDTH)

    # Write headers for the new block
    ws.cell(HEADER_ROW, insert_col, value=f"Ein {new_block_num}")
    ws.cell(HEADER_ROW, insert_col + 1, value=f"Aus {new_block_num}")
    ws.cell(HEADER_ROW, insert_col + 2, value=f"Stunden {new_block_num}")

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for c in range(insert_col, insert_col + BLOCK_WIDTH):
        cell = ws.cell(HEADER_ROW, c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")


def recalculate_day(ws: Worksheet, date: datetime.date, config: dict) -> None:
    """Recalculate Stunden, Gesamt, Saldo for a day row.

    Travel offset logic:
    - Read the day's Status (Home or Office).
    - If Office: add offset minutes to first Ein (subtract from start) and last Aus (add to end).
    - If Home: no offset.
    - When clocking in again (new block), the previous "last Aus" offset is removed
      because recalculate_day always recomputes from scratch.
    """
    row_num = find_day_row(ws, date)
    if row_num is None:
        return

    num_blocks = _count_blocks(ws)
    offset = cfg_mod.get_travel_offset(config)
    threshold = config.get("auto_break_threshold_hours", 6.0)
    deduction_min = config.get("auto_break_deduction_minutes", 30)

    # Read day status
    day_status = _read_day_status(ws, row_num)
    is_home = (day_status == "Home")

    # Read all blocks
    blocks = []
    for b in range(1, num_blocks + 1):
        ein_val = _to_time(ws.cell(row_num, _ein_col(b)).value)
        aus_val = _to_time(ws.cell(row_num, _aus_col(b)).value)
        blocks.append((b, ein_val, aus_val))

    # Travel offset: only for Office days
    # Applied to the very first Ein and the very last Aus
    first_ein_block = None
    last_aus_block = None
    if not is_home:
        for b, ein, aus in blocks:
            if ein is not None:
                first_ein_block = b
                break
        for b, ein, aus in reversed(blocks):
            if aus is not None:
                last_aus_block = b
                break

    total_work = 0.0

    for b, ein, aus in blocks:
        s_col = _stunden_col(b)
        if ein is None or aus is None:
            ws.cell(row_num, s_col, value=None)
            continue

        # Apply travel offset
        ein_dt = datetime.datetime.combine(date, ein)
        aus_dt = datetime.datetime.combine(date, aus)

        if b == first_ein_block:
            ein_dt -= datetime.timedelta(minutes=offset)
        if b == last_aus_block:
            aus_dt += datetime.timedelta(minutes=offset)

        diff = (aus_dt - ein_dt).total_seconds() / 3600.0
        if diff < 0:
            diff = 0.0

        hours = round(diff, 4)
        ws.cell(row_num, s_col, value=_hours_to_fraction(round(hours, 2)))
        ws.cell(row_num, s_col).number_format = HOURS_FMT
        ws.cell(row_num, s_col).alignment = Alignment(horizontal="center")

        total_work += hours

    # Apply auto break deduction
    if total_work > threshold + (deduction_min / 60.0):
        total_work -= deduction_min / 60.0
    elif total_work > threshold:
        total_work = threshold

    _, gesamt_col, soll_col, saldo_col = _find_summary_cols(ws)

    ws.cell(row_num, gesamt_col, value=_hours_to_fraction(round(total_work, 2)))
    ws.cell(row_num, gesamt_col).number_format = HOURS_FMT
    ws.cell(row_num, gesamt_col).alignment = Alignment(horizontal="center")

    soll_raw = ws.cell(row_num, soll_col).value
    if soll_raw is None:
        expected = cfg_mod.get_expected_hours(config, date.weekday())
        ws.cell(row_num, soll_col, value=_hours_to_fraction(expected))
        ws.cell(row_num, soll_col).number_format = HOURS_FMT
    else:
        expected = _read_hours_value(soll_raw)

    saldo = round(total_work - expected, 2)
    ws.cell(row_num, saldo_col, value=_hours_to_fraction(saldo))
    ws.cell(row_num, saldo_col).number_format = SALDO_FMT
    ws.cell(row_num, saldo_col).alignment = Alignment(horizontal="center")


def recalculate_sheet_summary(ws: Worksheet, year: int, month: int,
                              carry_over: float, config: dict) -> float:
    """Recalculate summary rows at the bottom. Returns cumulative balance."""
    _, gesamt_col, soll_col, saldo_col = _find_summary_cols(ws)

    # Find summary rows by label
    sum_row = None
    carry_row = None
    cum_row = None
    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        val = ws.cell(row, 1).value
        if val == "Summe":
            sum_row = row
        elif val == "Uebertrag":
            carry_row = row
        elif val == "Kumuliert":
            cum_row = row

    if sum_row is None:
        return carry_over

    # Sum over data rows (values are stored as time fractions, read as timedelta)
    total_gesamt = 0.0
    total_soll = 0.0
    total_saldo = 0.0
    for row in range(HEADER_ROW + 1, sum_row):
        # Skip blank separator rows
        if ws.cell(row, 1).value is None:
            continue
        g = ws.cell(row, gesamt_col).value
        s = ws.cell(row, soll_col).value
        b = ws.cell(row, saldo_col).value
        if g is not None:
            total_gesamt += _read_hours_value(g)
        if s is not None:
            total_soll += _read_hours_value(s)
        if b is not None:
            total_saldo += _read_hours_value(b)

    # totals are now in decimal hours (via _read_hours_value), convert back to fractions
    ws.cell(sum_row, gesamt_col, value=_hours_to_fraction(round(total_gesamt, 2)))
    ws.cell(sum_row, gesamt_col).number_format = HOURS_FMT
    ws.cell(sum_row, soll_col, value=_hours_to_fraction(round(total_soll, 2)))
    ws.cell(sum_row, soll_col).number_format = HOURS_FMT
    ws.cell(sum_row, saldo_col, value=_hours_to_fraction(round(total_saldo, 2)))
    ws.cell(sum_row, saldo_col).number_format = SALDO_FMT

    if carry_row:
        ws.cell(carry_row, saldo_col, value=_hours_to_fraction(round(carry_over, 2)))
        ws.cell(carry_row, saldo_col).number_format = SALDO_FMT

    cumulative = round(carry_over + total_saldo, 2)
    if cum_row:
        ws.cell(cum_row, saldo_col, value=_hours_to_fraction(cumulative))
        ws.cell(cum_row, saldo_col).number_format = SALDO_FMT

    return cumulative


def fill_missing_days(ws: Worksheet, config: dict, up_to_date: datetime.date) -> None:
    """For workdays with no stamps before up_to_date, set total = expected."""
    _, gesamt_col, soll_col, saldo_col = _find_summary_cols(ws)

    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        date_val = ws.cell(row, 1).value
        if date_val is None:
            continue

        # Parse date
        if isinstance(date_val, str):
            try:
                dt = datetime.datetime.strptime(date_val, "%d.%m.%Y").date()
            except ValueError:
                continue
        elif isinstance(date_val, datetime.datetime):
            dt = date_val.date()
        elif isinstance(date_val, datetime.date):
            dt = date_val
        else:
            continue

        if dt >= up_to_date:
            continue

        # Check if any stamps exist
        num_blocks = _count_blocks(ws)
        has_stamps = False
        for b in range(1, num_blocks + 1):
            if ws.cell(row, _ein_col(b)).value is not None:
                has_stamps = True
                break

        if has_stamps:
            continue

        # Check if already filled
        if ws.cell(row, gesamt_col).value is not None:
            continue

        expected = cfg_mod.get_expected_hours(config, dt.weekday())
        ws.cell(row, gesamt_col, value=_hours_to_fraction(expected))
        ws.cell(row, gesamt_col).number_format = HOURS_FMT
        ws.cell(row, gesamt_col).alignment = Alignment(horizontal="center")
        ws.cell(row, soll_col, value=_hours_to_fraction(expected))
        ws.cell(row, soll_col).number_format = HOURS_FMT
        ws.cell(row, saldo_col, value=0.0)
        ws.cell(row, saldo_col).number_format = SALDO_FMT
        ws.cell(row, saldo_col).alignment = Alignment(horizontal="center")


def iter_day_rows_with_data(ws: Worksheet) -> List[Tuple[datetime.date, Optional[float], float, Optional[float]]]:
    """Iterate over data rows, returning (date, total, expected, saldo) tuples."""
    _, gesamt_col, soll_col, saldo_col = _find_summary_cols(ws)
    results = []

    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        date_val = ws.cell(row, 1).value
        if date_val is None:
            continue
        # Stop at summary rows
        if isinstance(date_val, str) and date_val in ("Summe", "Uebertrag", "Kumuliert"):
            break

        if isinstance(date_val, str):
            try:
                dt = datetime.datetime.strptime(date_val, "%d.%m.%Y").date()
            except ValueError:
                continue
        elif isinstance(date_val, datetime.datetime):
            dt = date_val.date()
        elif isinstance(date_val, datetime.date):
            dt = date_val
        else:
            continue

        total = ws.cell(row, gesamt_col).value
        expected = ws.cell(row, soll_col).value
        saldo = ws.cell(row, saldo_col).value

        total_f = _read_hours_value(total) if total is not None else None
        expected_f = _read_hours_value(expected) if expected is not None else 0.0
        saldo_f = _read_hours_value(saldo) if saldo is not None else None

        results.append((dt, total_f, expected_f, saldo_f))

    return results


def has_open_stamp(ws: Worksheet, date: datetime.date) -> bool:
    """Check if a day has an Ein without a matching Aus."""
    row_num = find_day_row(ws, date)
    if row_num is None:
        return False

    num_blocks = _count_blocks(ws)
    for b in range(1, num_blocks + 1):
        ein_val = ws.cell(row_num, _ein_col(b)).value
        aus_val = ws.cell(row_num, _aus_col(b)).value
        if ein_val is not None and aus_val is None:
            return True
    return False


def set_day_sick(ws: Worksheet, date: datetime.date, config: dict) -> None:
    """Set a day's Gesamt to its Soll (expected hours), resulting in zero saldo.
    Used when clocking out sick -- you shouldn't lose hours for being ill."""
    row_num = find_day_row(ws, date)
    if row_num is None:
        return

    status_col, gesamt_col, soll_col, saldo_col = _find_summary_cols(ws)
    ws.cell(row_num, status_col, value="Krank")
    ws.cell(row_num, status_col).alignment = Alignment(horizontal="center")

    expected = ws.cell(row_num, soll_col).value
    if expected is None:
        expected = cfg_mod.get_expected_hours(config, date.weekday())

    expected_hours = _read_hours_value(expected)
    ws.cell(row_num, gesamt_col, value=_hours_to_fraction(expected_hours))
    ws.cell(row_num, gesamt_col).number_format = HOURS_FMT
    ws.cell(row_num, gesamt_col).alignment = Alignment(horizontal="center")
    ws.cell(row_num, saldo_col, value=0.0)
    ws.cell(row_num, saldo_col).number_format = SALDO_FMT
    ws.cell(row_num, saldo_col).alignment = Alignment(horizontal="center")


def calculate_current_hours(ws: Worksheet, date: datetime.date, config: dict) -> Optional[float]:
    """Calculate hours worked so far today, treating 'now' as a virtual clock-out
    for any open stamp block. Returns None if no stamps exist."""
    row_num = find_day_row(ws, date)
    if row_num is None:
        return None

    num_blocks = _count_blocks(ws)
    offset = cfg_mod.get_travel_offset(config)
    threshold = config.get("auto_break_threshold_hours", 6.0)
    deduction_min = config.get("auto_break_deduction_minutes", 30)

    day_status = _read_day_status(ws, row_num)
    is_home = (day_status == "Home")

    now = datetime.datetime.now().time().replace(second=0, microsecond=0)

    blocks = []
    has_any = False
    for b in range(1, num_blocks + 1):
        ein_val = _to_time(ws.cell(row_num, _ein_col(b)).value)
        aus_val = _to_time(ws.cell(row_num, _aus_col(b)).value)
        if ein_val is not None:
            has_any = True
        # If open stamp, treat now as virtual aus
        if ein_val is not None and aus_val is None:
            aus_val = now
        blocks.append((b, ein_val, aus_val))

    if not has_any:
        return None

    # Determine offset blocks (same logic as recalculate_day)
    first_ein_block = None
    last_aus_block = None
    if not is_home:
        for b, ein, aus in blocks:
            if ein is not None:
                first_ein_block = b
                break
        for b, ein, aus in reversed(blocks):
            if aus is not None:
                last_aus_block = b
                break

    total_work = 0.0
    for b, ein, aus in blocks:
        if ein is None or aus is None:
            continue
        ein_dt = datetime.datetime.combine(date, ein)
        aus_dt = datetime.datetime.combine(date, aus)

        if b == first_ein_block:
            ein_dt -= datetime.timedelta(minutes=offset)
        if b == last_aus_block:
            aus_dt += datetime.timedelta(minutes=offset)

        diff = (aus_dt - ein_dt).total_seconds() / 3600.0
        if diff < 0:
            diff = 0.0
        total_work += diff

    # Apply auto break deduction
    if total_work > threshold + (deduction_min / 60.0):
        total_work -= deduction_min / 60.0
    elif total_work > threshold:
        total_work = threshold

    return round(total_work, 2)


def save_workbook(wb: Workbook, data_dir: Path, year: int) -> None:
    path = get_workbook_path(data_dir, year)
    wb.save(path)
